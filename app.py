from flask import Flask, render_template, request, jsonify, redirect, url_for
from openpyxl import Workbook, load_workbook
import os, time, hashlib

app = Flask(__name__)
generated_tokens = {}
used_tokens = set()
parties = ["Party A", "Party B", "Party C"]

# Excel setup
excel_file = "votes.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "votes"
    ws.append(["Voter Name", "Token", "Party", "Timestamp"])
    wb.save(excel_file)

def generate_token(name):
    raw = name + str(time.time())
    return hashlib.sha256(raw.encode()).hexdigest()[:8]

@app.route("/", methods=["GET", "POST"])
def index():
    message = ""
    message_type = "success"

    if request.method == "POST":
        action = request.form.get("action")

        if action == "register":
            name = request.form.get("name").strip().upper()
            wb = load_workbook(excel_file)
            ws = wb.active
            existing_names = [row[0].value.strip().upper() for row in ws.iter_rows(min_row=2) if row[0].value]

            if name in generated_tokens or name in existing_names:
                message = "Name already registered. Please use a different name."
                message_type = "error"
            else:
                token = generate_token(name)
                generated_tokens[name] = token
                message = f"Token generated: {token}. Save this token to vote."
                message_type = "success"

        elif action == "vote":
            token = request.form.get("token").strip()
            vote = request.form.get("vote")

            if token not in generated_tokens.values():
                message = "Invalid token."
                message_type = "error"
            elif token in used_tokens:
                message = "This token has already been used."
                message_type = "error"
            elif vote not in parties:
                message = "Invalid party selected."
                message_type = "error"
            else:
                used_tokens.add(token)
                name = [n for n, t in generated_tokens.items() if t == token][0]
                wb = load_workbook(excel_file)
                ws = wb.active
                ws.append([name, token, vote, time.strftime("%Y-%m-%d %H:%M:%S")])
                wb.save(excel_file)
                return redirect(url_for("results_page"))  # 🚀 redirect to result page


    return render_template("index.html", message=message, parties=parties, message_type=message_type)

@app.route("/results-page")
def results_page():
    return render_template("results.html", parties=parties)

@app.route("/results")
def results():
    wb = load_workbook(excel_file)
    ws = wb.active

    counts = {party: 0 for party in parties}
    for row in ws.iter_rows(min_row=2, values_only=True):
        party = row[2]
        if party in counts:
            counts[party] += 1

    return jsonify(counts)

if __name__ == "__main__":
    app.run(debug=True)
